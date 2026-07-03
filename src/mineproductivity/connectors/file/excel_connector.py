"""``ExcelConnector``: reads the identical row shape as
:class:`~mineproductivity.connectors.file.csv_connector.CSVConnector`
from an ``.xlsx`` workbook instead of a CSV file.
"""

from __future__ import annotations

import logging
from collections.abc import Callable, Iterator
from datetime import datetime, timezone
from pathlib import Path
from typing import ClassVar, TypeVar

from mineproductivity.connectors.base import FMSConnector, IngestionMode
from mineproductivity.connectors.exceptions import MappingError, SourceUnavailableError
from mineproductivity.connectors.file._common import FileRowNormalizer, parse_source_datetime
from mineproductivity.connectors.health import ConnectorHealth, HealthStatus
from mineproductivity.connectors.normalization import Normalizer
from mineproductivity.events import CycleEvent, DelayEvent

__all__ = ["ExcelConnector"]

_logger = logging.getLogger(__name__)
_TEvent = TypeVar("_TEvent")


class ExcelConnector(FMSConnector):
    """Reads a haul-cycle (and, optionally, delay) ``.xlsx`` export,
    using the identical row shape and normalization logic as
    :class:`~mineproductivity.connectors.file.csv_connector.CSVConnector`
    (first sheet only; first row is the header row).

    Requires the optional ``openpyxl`` dependency (the ``connectors``
    extra) -- imported lazily inside :meth:`_read`, never at module
    import time, so importing this module (or ``mineproductivity.connectors``
    itself) never requires ``openpyxl`` to be installed unless an
    ``ExcelConnector`` is actually instantiated and read from.

    Not thread-safe for concurrent calls against the same instance, for
    the same reason as ``CSVConnector`` (design spec §24).
    """

    name: ClassVar[str] = "excel"
    supported_modes: ClassVar[tuple[IngestionMode, ...]] = (IngestionMode.BATCH,)

    def __init__(
        self,
        path: str | Path,
        shift_id: str,
        *,
        delay_path: str | Path | None = None,
        source_timezone: str = "UTC",
        normalizer: Normalizer | None = None,
    ) -> None:
        self._path = Path(path)
        self._delay_path = Path(delay_path) if delay_path is not None else None
        self._shift_id = shift_id
        self._source_timezone = source_timezone
        self._normalizer: Normalizer = normalizer if normalizer is not None else FileRowNormalizer()
        self._last_successful_pull_utc: datetime | None = None
        self._last_status: HealthStatus = HealthStatus.UNKNOWN

    def get_cycle_data(self, since: datetime, until: datetime) -> Iterator[CycleEvent]:
        yield from self._read(self._path, since, until, self._normalizer.normalize_cycle)

    def get_delay_data(self, since: datetime, until: datetime) -> Iterator[DelayEvent]:
        if self._delay_path is None:
            return
        yield from self._read(self._delay_path, since, until, self._normalizer.normalize_delay)

    def _iter_rows(self, path: Path) -> Iterator[dict[str, object]]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - exercised only without the extra installed
            raise SourceUnavailableError(
                "ExcelConnector requires the optional 'openpyxl' dependency "
                "(install the 'connectors' extra: pip install 'mineproductivity[connectors]')"
            ) from exc

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(values_only=True)
            header = [str(cell) if cell is not None else "" for cell in next(rows)]
            for values in rows:
                yield dict(zip(header, values, strict=False))
        finally:
            workbook.close()

    def _read(
        self,
        path: Path,
        since: datetime,
        until: datetime,
        normalize: Callable[[dict[str, object]], _TEvent],
    ) -> Iterator[_TEvent]:
        if not path.exists():
            self._last_status = HealthStatus.UNHEALTHY
            raise SourceUnavailableError(f"Excel source not found: {path}")
        try:
            self._last_status = HealthStatus.HEALTHY
            self._last_successful_pull_utc = datetime.now(tz=timezone.utc)
            for row in self._iter_rows(path):
                raw_time = row.get("event_time")
                if not raw_time:
                    _logger.warning("row missing event_time, skipped: %r", row)
                    continue
                try:
                    event_time = parse_source_datetime(
                        str(raw_time), source_timezone=self._source_timezone
                    )
                except MappingError as exc:
                    _logger.warning("unparseable event_time, row skipped: %s", exc)
                    continue
                if not (since <= event_time < until):
                    continue
                raw = {**row, "shift_id": self._shift_id}
                try:
                    yield normalize(raw)
                except MappingError as exc:
                    _logger.warning("row failed to normalize and was skipped: %s", exc)
                    continue
        except OSError as exc:
            self._last_status = HealthStatus.UNHEALTHY
            raise SourceUnavailableError(f"failed to read {path}: {exc}") from exc

    def health_check(self) -> ConnectorHealth:
        return ConnectorHealth(
            status=self._last_status,
            last_successful_pull_utc=self._last_successful_pull_utc,
        )
